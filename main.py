import os
import uuid
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from models import db, Card, Lot, Folder, PriceHistory
from search import search_card_price, search_ebay_sold
from PIL import Image
import anthropic
import base64
import json

main = Blueprint('main', __name__)

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def resize_image(filepath, max_size=(800, 800)):
    img = Image.open(filepath)
    try:
        exif = img._getexif()
        if exif:
            orientation_key = 274
            if orientation_key in exif:
                orientation = exif[orientation_key]
                rotations = {3: 180, 6: 270, 8: 90}
                if orientation in rotations:
                    img = img.rotate(rotations[orientation], expand=True)
    except Exception:
        pass
    img.thumbnail(max_size, Image.LANCZOS)
    img.save(filepath, optimize=True, quality=85)

def build_ai_prompt(card, ebay_data):
    """Build a prompt for Claude to analyze card's investment potential."""
    listings_text = "\n".join([
        f"- {l['title']}: ${l['price']:.2f} ({l.get('condition', 'Unknown')})"
        for l in ebay_data.get('listings', [])[:5]
    ])

    profit = card.profit_loss if hasattr(card, 'profit_loss') else (card.sell_price - card.buy_price if card.sell_price else None)
    profit_text = f"${profit:.2f}" if profit is not None else "Not sold yet"
    market_price_text = f"${card.market_price:.2f}" if card.market_price else "Unknown"
    sell_price_text = f"${card.sell_price:.2f}" if card.sell_price else "Not set"

    prompt = f"""You are a sports card market analyst. Analyze this specific card in the user's collection and provide a KEEP vs SELL recommendation.

CARD DETAILS:
- Player: {card.player_name}
- Sport: {card.sport}
- Year: {card.year}
- Manufacturer: {card.manufacturer}
- Condition: {card.condition}
- Buy Price: ${card.buy_price:.2f}
- Current Market Price: {market_price_text}
- Current Sell Price: {sell_price_text}
- Profit/Loss: {profit_text}

EBAY MARKET DATA (Recent Sold Listings):
- Average Price: ${ebay_data['average']:.2f}
- Low: ${ebay_data['low']:.2f}
- High: ${ebay_data['high']:.2f}
- Number of Recent Sales: {ebay_data['count']}

Sample Recent Sales:
{listings_text}

Based on the card's acquisition cost, current market conditions, and recent sales, provide:
1. A one-word recommendation: KEEP, SELL, or HOLD
2. A 2-3 sentence analysis explaining whether they should hold this specific card for appreciation, sell it now for current value, or hold & monitor.

Format as JSON:
{{"recommendation": "KEEP/SELL/HOLD", "analysis": "Your analysis here"}}"""

    return prompt

@main.route('/')
@main.route('/dashboard')
@login_required
def dashboard():
    sport_filter = request.args.get('sport', 'all')
    sort_by = request.args.get('sort', 'date')
    folder_filter = request.args.get('folder', 'all')
    view_mode = request.args.get('view', 'all')

    query = Card.query.filter_by(user_id=current_user.id)

    if sport_filter != 'all':
        query = query.filter_by(sport=sport_filter)

    if folder_filter != 'all':
        query = query.filter_by(folder_id=int(folder_filter))

    if sort_by == 'value_high':
        cards = sorted(query.all(), key=lambda c: c.market_price or c.sell_price or c.buy_price, reverse=True)
    elif sort_by == 'value_low':
        cards = sorted(query.all(), key=lambda c: c.market_price or c.sell_price or c.buy_price)
    elif sort_by == 'profit_high':
        cards = sorted(query.all(), key=lambda c: c.profit_loss or 0, reverse=True)
    elif sort_by == 'profit_low':
        cards = sorted(query.all(), key=lambda c: c.profit_loss or 0)
    elif sort_by == 'name':
        cards = sorted(query.all(), key=lambda c: c.player_name)
    else:
        cards = query.order_by(Card.date_added.desc()).all()

    all_cards = Card.query.filter_by(user_id=current_user.id).all()
    total_invested = sum(card.buy_price for card in all_cards)

    est_market_value = sum(
        card.market_price or card.buy_price
        for card in all_cards
        if not card.sell_price
    )

    realized_gains = sum(
        card.sell_price - card.buy_price
        for card in all_cards
        if card.sell_price
    )

    total_value = sum(
        card.sell_price if card.sell_price
        else (card.market_price or card.buy_price)
        for card in all_cards
    )

    total_profit_loss = total_value - total_invested
    sports = ['Baseball', 'Basketball', 'Football', 'Hockey', 'Soccer', 'Other']
    folders = Folder.query.filter_by(user_id=current_user.id).order_by(Folder.name).all()

    unassigned_cards = Card.query.filter_by(
        user_id=current_user.id, folder_id=None
    ).order_by(Card.date_added.desc()).all()

    from collections import defaultdict
    sport_counts = defaultdict(int)
    sport_invested = defaultdict(float)
    sport_value = defaultdict(float)
    for card in all_cards:
        sport = card.sport or "Unknown"
        sport_counts[sport] += 1
        sport_invested[sport] += card.buy_price
        sport_value[sport] += card.sell_price if card.sell_price else (card.market_price or card.buy_price)
    chart_sport_labels = list(sport_counts.keys())
    chart_sport_counts = [sport_counts[s] for s in chart_sport_labels]
    chart_sport_invested = [round(sport_invested[s], 2) for s in chart_sport_labels]
    chart_sport_value = [round(sport_value[s], 2) for s in chart_sport_labels]

    return render_template('dashboard.html',
        cards=cards,
        total_invested=total_invested,
        total_value=total_value,
        total_profit_loss=total_profit_loss,
        est_market_value=est_market_value,
        realized_gains=realized_gains,
        sport_filter=sport_filter,
        sort_by=sort_by,
        sports=sports,
        folders=folders,
        folder_filter=folder_filter,
        view_mode=view_mode,
        unassigned_cards=unassigned_cards,
        chart_sport_labels=chart_sport_labels,
        chart_sport_counts=chart_sport_counts,
        chart_sport_invested=chart_sport_invested,
        chart_sport_value=chart_sport_value
    )

@main.route('/add_card', methods=['GET', 'POST'])
@login_required
def add_card():
    if current_user.is_demo:
        flash('Demo account cannot make changes.')
        return redirect(url_for('main.dashboard'))
    folders = Folder.query.filter_by(user_id=current_user.id).order_by(Folder.name).all()
    if request.method == 'POST':
        player_name = request.form.get('player_name')
        sport = request.form.get('sport')
        year = request.form.get('year')
        manufacturer = request.form.get('manufacturer')
        condition = request.form.get('condition')
        buy_price = request.form.get('buy_price')
        sell_price = request.form.get('sell_price')
        notes = request.form.get('notes')
        folder_id = request.form.get('folder_id')

        photo_filename = None
        if 'photo' in request.files:
            photo = request.files['photo']
            if photo and allowed_file(photo.filename):
                filename = str(uuid.uuid4()) + '_' + secure_filename(photo.filename)
                upload_folder = 'static/uploads'
                os.makedirs(upload_folder, exist_ok=True)
                filepath = os.path.join(upload_folder, filename)
                photo.save(filepath)
                resize_image(filepath)
                photo_filename = filename

        market_price = None
        if player_name:
            price, error = search_card_price(
                player_name, year, manufacturer, sport, condition
            )
            if price:
                market_price = price

        new_card = Card(
            player_name=player_name,
            sport=sport,
            year=year,
            manufacturer=manufacturer,
            condition=condition,
            buy_price=float(buy_price),
            sell_price=float(sell_price) if sell_price else None,
            photo_filename=photo_filename,
            market_price=market_price,
            notes=notes,
            folder_id=int(folder_id) if folder_id else None,
            user_id=current_user.id
        )
        db.session.add(new_card)
        db.session.commit()

        if market_price:
            history = PriceHistory(card_id=new_card.id, price=market_price)
            db.session.add(history)
            db.session.commit()
            flash(f'Card added! Market price found: ${market_price:.2f}')
        else:
            flash('Card added! No market price found automatically.')

        return redirect(url_for('main.dashboard'))

    return render_template('add_card.html', folders=folders)

@main.route('/edit_card/<int:card_id>', methods=['GET', 'POST'])
@login_required
def edit_card(card_id):
    if current_user.is_demo:
        flash('Demo account cannot make changes.')
        return redirect(url_for('main.dashboard'))
    card = Card.query.get_or_404(card_id)
    folders = Folder.query.filter_by(user_id=current_user.id).order_by(Folder.name).all()
    if card.user_id != current_user.id:
        flash('Permission denied.')
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':
        card.player_name = request.form.get('player_name')
        card.sport = request.form.get('sport')
        card.year = request.form.get('year')
        card.manufacturer = request.form.get('manufacturer')
        card.condition = request.form.get('condition')
        card.buy_price = float(request.form.get('buy_price'))
        sell_price = request.form.get('sell_price')
        card.sell_price = float(sell_price) if sell_price else None
        card.notes = request.form.get('notes')
        folder_id = request.form.get('folder_id')
        card.folder_id = int(folder_id) if folder_id else None

        if 'photo' in request.files:
            photo = request.files['photo']
            if photo and allowed_file(photo.filename):
                filename = str(uuid.uuid4()) + '_' + secure_filename(photo.filename)
                upload_folder = 'static/uploads'
                os.makedirs(upload_folder, exist_ok=True)
                filepath = os.path.join(upload_folder, filename)
                photo.save(filepath)
                resize_image(filepath)
                card.photo_filename = filename

        db.session.commit()
        flash('Card updated successfully.')
        return redirect(url_for('main.card_detail', card_id=card.id))

    return render_template('edit_card.html', card=card, folders=folders)

@main.route('/card/<int:card_id>')
@login_required
def card_detail(card_id):
    card = Card.query.get_or_404(card_id)
    if card.user_id != current_user.id:
        flash('Permission denied.')
        return redirect(url_for('main.dashboard'))
    price_history = PriceHistory.query.filter_by(card_id=card_id).order_by(PriceHistory.date_recorded).all()
    ebay_data, ebay_error = search_ebay_sold(
        card.player_name, card.year, card.manufacturer, card.sport, card.condition
    )
    return render_template('card_detail.html', card=card, price_history=price_history,
                           ebay_data=ebay_data, ebay_error=ebay_error)

@main.route('/search_price/<int:card_id>')
@login_required
def search_price(card_id):
    card = Card.query.get_or_404(card_id)
    if card.user_id != current_user.id:
        flash('Permission denied.')
        return redirect(url_for('main.dashboard'))

    price, error = search_card_price(
        card.player_name, card.year, card.manufacturer, card.sport, card.condition
    )

    if price:
        card.market_price = price
        db.session.commit()
        history = PriceHistory(card_id=card.id, price=price)
        db.session.add(history)
        db.session.commit()
        flash(f'Market price updated to ${price:.2f}')
    else:
        flash('Could not find market price. Try again later.')

    return redirect(url_for('main.dashboard'))

@main.route('/export_csv')
@login_required
def export_csv():
    import csv
    import io
    from flask import Response

    cards = Card.query.filter_by(user_id=current_user.id).all()

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        'Player Name', 'Sport', 'Year', 'Manufacturer',
        'Condition', 'Buy Price', 'Sell Price', 'Market Price',
        'Profit/Loss', 'Date Added'
    ])

    for card in cards:
        writer.writerow([
            card.player_name,
            card.sport or '',
            card.year or '',
            card.manufacturer or '',
            card.condition or '',
            f'${card.buy_price:.2f}',
            f'${card.sell_price:.2f}' if card.sell_price else '',
            f'${card.market_price:.2f}' if card.market_price else '',
            f'${card.profit_loss:.2f}' if card.profit_loss is not None else '',
            card.date_added.strftime('%Y-%m-%d') if card.date_added else ''
        ])

    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=card-collection.csv'}
    )

@main.route('/import_cards', methods=['GET', 'POST'])
@login_required
def import_cards():
    import csv
    import io
    import openpyxl

    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected.')
            return redirect(url_for('main.import_cards'))

        file = request.files['file']
        if file.filename == '':
            flash('No file selected.')
            return redirect(url_for('main.import_cards'))

        filename = file.filename.lower()
        rows = []
        errors = []
        imported = 0

        try:
            if filename.endswith('.csv'):
                stream = io.StringIO(file.stream.read().decode('utf-8'))
                reader = csv.DictReader(stream)
                rows = list(reader)

            elif filename.endswith('.xlsx'):
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))

            else:
                flash('Please upload a CSV or Excel (.xlsx) file.')
                return redirect(url_for('main.import_cards'))

        except Exception as e:
            flash(f'Error reading file: {str(e)}')
            return redirect(url_for('main.import_cards'))

        for i, row in enumerate(rows, start=2):
            try:
                player_name = row.get('Player Name') or row.get('player_name') or row.get('Name')
                if not player_name:
                    errors.append(f'Row {i}: Missing player name')
                    continue

                buy_price_raw = row.get('Buy Price') or row.get('buy_price') or '0'
                buy_price_str = str(buy_price_raw).replace('$', '').replace(',', '').strip()
                buy_price = float(buy_price_str) if buy_price_str else 0.0

                sell_price_raw = row.get('Sell Price') or row.get('sell_price') or ''
                sell_price_str = str(sell_price_raw).replace('$', '').replace(',', '').strip()
                sell_price = float(sell_price_str) if sell_price_str else None

                market_price_raw = row.get('Market Price') or row.get('market_price') or ''
                market_price_str = str(market_price_raw).replace('$', '').replace(',', '').strip()
                market_price = float(market_price_str) if market_price_str else None

                new_card = Card(
                    player_name=str(player_name).strip(),
                    sport=str(row.get('Sport') or row.get('sport') or '').strip() or None,
                    year=str(row.get('Year') or row.get('year') or '').strip() or None,
                    manufacturer=str(row.get('Manufacturer') or row.get('manufacturer') or '').strip() or None,
                    condition=str(row.get('Condition') or row.get('condition') or '').strip() or None,
                    buy_price=buy_price,
                    sell_price=sell_price,
                    market_price=market_price,
                    user_id=current_user.id
                )
                db.session.add(new_card)
                imported += 1

            except Exception as e:
                errors.append(f'Row {i}: {str(e)}')

        db.session.commit()

        if errors:
            flash(f'Imported {imported} cards. {len(errors)} rows had errors: {", ".join(errors[:3])}')
        else:
            flash(f'Successfully imported {imported} cards.')

        return redirect(url_for('main.dashboard'))

    return render_template('import_cards.html')

@main.route('/lots')
@login_required
def lots():
    lots = Lot.query.filter_by(user_id=current_user.id).order_by(Lot.date_sold.desc()).all()
    return render_template('lots.html', lots=lots)

@main.route('/create_lot', methods=['GET', 'POST'])
@login_required
def create_lot():
    if current_user.is_demo:
        flash('Demo account cannot make changes.')
        return redirect(url_for('main.dashboard'))
    if request.method == 'POST':
        name = request.form.get('name')
        total_sale_price = float(request.form.get('total_sale_price'))
        notes = request.form.get('notes')
        card_ids = request.form.getlist('card_ids')

        if not card_ids:
            flash('Please select at least one card for the lot.')
            return redirect(url_for('main.create_lot'))

        new_lot = Lot(
            name=name,
            total_sale_price=total_sale_price,
            notes=notes,
            user_id=current_user.id
        )
        db.session.add(new_lot)
        db.session.flush()

        price_per_card = total_sale_price / len(card_ids)

        for card_id in card_ids:
            card = Card.query.get(int(card_id))
            if card and card.user_id == current_user.id:
                card.lot_id = new_lot.id
                card.sell_price = round(price_per_card, 2)

        db.session.commit()
        flash(f'Lot created with {len(card_ids)} cards sold for ${total_sale_price:.2f}')
        return redirect(url_for('main.lots'))

    unsold_cards = Card.query.filter_by(
        user_id=current_user.id,
        lot_id=None
    ).filter(Card.sell_price == None).order_by(Card.player_name).all()

    return render_template('create_lot.html', cards=unsold_cards)

@main.route('/delete_lot/<int:lot_id>')
@login_required
def delete_lot(lot_id):
    if current_user.is_demo:
        flash('Demo account cannot make changes.')
        return redirect(url_for('main.dashboard'))
    lot = Lot.query.get_or_404(lot_id)
    if lot.user_id != current_user.id:
        flash('Permission denied.')
        return redirect(url_for('main.lots'))

    for card in lot.cards:
        card.lot_id = None
        card.sell_price = None

    db.session.delete(lot)
    db.session.commit()
    flash('Lot deleted and cards returned to collection.')
    return redirect(url_for('main.lots'))

@main.route('/delete_card/<int:card_id>')
@login_required
def delete_card(card_id):
    if current_user.is_demo:
        flash('Demo account cannot make changes.')
        return redirect(url_for('main.dashboard'))
    card = Card.query.get_or_404(card_id)
    if card.user_id != current_user.id:
        flash('Permission denied.')
        return redirect(url_for('main.dashboard'))
    PriceHistory.query.filter_by(card_id=card.id).delete()
    db.session.delete(card)
    db.session.commit()
    flash('Card deleted.')
    return redirect(url_for('main.dashboard'))

@main.route('/bulk_assign_folder', methods=['POST'])
@login_required
def bulk_assign_folder():
    card_ids = request.form.getlist('card_ids')
    folder_id = request.form.get('folder_id')

    if not card_ids:
        flash('No cards selected.')
        return redirect(url_for('main.dashboard'))

    folder_id = int(folder_id) if folder_id else None

    for card_id in card_ids:
        card = Card.query.get(int(card_id))
        if card and card.user_id == current_user.id:
            card.folder_id = folder_id

    db.session.commit()

    if folder_id:
        folder = Folder.query.get(folder_id)
        flash(f'Moved {len(card_ids)} cards to "{folder.name}".')
    else:
        flash(f'Removed {len(card_ids)} cards from their folder.')

    return redirect(url_for('main.dashboard'))

@main.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    from werkzeug.security import generate_password_hash, check_password_hash
    if request.method == 'POST':
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        security_question = request.form.get('security_question')
        security_answer = request.form.get('security_answer', '').strip()

        if not check_password_hash(current_user.password, current_password):
            flash('Current password is incorrect.')
            return redirect(url_for('main.change_password'))

        if new_password != confirm_password:
            flash('New passwords do not match.')
            return redirect(url_for('main.change_password'))

        current_user.password = generate_password_hash(new_password)

        if security_question:
            current_user.security_question = security_question
        if security_answer:
            current_user.security_answer = security_answer.lower().strip()

        db.session.commit()
        flash('Password and security question updated successfully.')
        return redirect(url_for('main.dashboard'))

    return render_template('change_password.html')

# --- Folder routes ---

@main.route('/create_folder', methods=['POST'])
@login_required
def create_folder():
    if current_user.is_demo:
        flash('Demo account cannot make changes.')
        return redirect(url_for('main.dashboard'))
    name = request.form.get('name', '').strip()
    if not name:
        flash('Folder name cannot be empty.')
        return redirect(url_for('main.dashboard'))
    existing = Folder.query.filter_by(user_id=current_user.id, name=name).first()
    if existing:
        flash('A folder with that name already exists.')
        return redirect(url_for('main.dashboard'))
    folder = Folder(name=name, user_id=current_user.id)
    db.session.add(folder)
    db.session.commit()
    flash(f'Folder "{name}" created.')
    return redirect(url_for('main.dashboard'))

@main.route('/delete_folder/<int:folder_id>')
@login_required
def delete_folder(folder_id):
    if current_user.is_demo:
        flash('Demo account cannot make changes.')
        return redirect(url_for('main.dashboard'))
    folder = Folder.query.get_or_404(folder_id)
    if folder.user_id != current_user.id:
        flash('Permission denied.')
        return redirect(url_for('main.dashboard'))
    for card in folder.cards:
        card.folder_id = None
    db.session.delete(folder)
    db.session.commit()
    flash(f'Folder "{folder.name}" deleted. Cards have been moved back to your collection.')
    return redirect(url_for('main.dashboard'))

@main.route('/rename_folder/<int:folder_id>', methods=['POST'])
@login_required
def rename_folder(folder_id):
    if current_user.is_demo:
        flash('Demo account cannot make changes.')
        return redirect(url_for('main.dashboard'))
    folder = Folder.query.get_or_404(folder_id)
    if folder.user_id != current_user.id:
        flash('Permission denied.')
        return redirect(url_for('main.dashboard'))
    new_name = request.form.get('name', '').strip()
    if not new_name:
        flash('Folder name cannot be empty.')
        return redirect(url_for('main.dashboard'))
    folder.name = new_name
    db.session.commit()
    flash(f'Folder renamed to "{new_name}".')
    return redirect(url_for('main.dashboard'))

@main.route('/scan_card', methods=['POST'])
@login_required
def scan_card():
    if 'image' not in request.files:
        return jsonify({'error': 'No image provided'}), 400

    image = request.files['image']
    image_data = base64.b64encode(image.read()).decode('utf-8')
    media_type = image.content_type

    try:
        client = anthropic.Anthropic(api_key=os.getenv('ANTHROPIC_API_KEY'))
        message = client.messages.create(
            model='claude-opus-4-5',
            max_tokens=1000,
            messages=[{
                'role': 'user',
                'content': [
                    {
                        'type': 'image',
                        'source': {
                            'type': 'base64',
                            'media_type': media_type,
                            'data': image_data
                        }
                    },
                    {
                        'type': 'text',
                        'text': '''Analyze this sports card image and extract the following information. Respond ONLY with a valid JSON object, no markdown, no extra text.

{
  "player_name": "full player name or empty string if not found",
  "sport": "one of: Baseball, Basketball, Football, Hockey, Soccer, Other, or empty string",
  "year": "4 digit year or empty string if not found",
  "manufacturer": "card manufacturer/brand like Topps, Panini, Upper Deck, Bowman, Fleer, Donruss, Score, etc. or empty string",
  "notes": "any other useful details like card set name, card number, rookie card, special edition, etc. or empty string"
}'''
                    }
                ]
            }]
        )
        result = json.loads(message.content[0].text.strip())
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@main.route('/analytics')
@login_required
def analytics():
    from collections import defaultdict
    all_cards = Card.query.filter_by(user_id=current_user.id).all()

    total_cards = len(all_cards)
    total_invested = sum(c.buy_price for c in all_cards)
    cards_sold = sum(1 for c in all_cards if c.sell_price)

    realized_gains = sum(
        c.sell_price - c.buy_price for c in all_cards if c.sell_price
    )
    unrealized = sum(
        (c.market_price or c.buy_price) - c.buy_price
        for c in all_cards if not c.sell_price
    )
    est_market_value = sum(
        c.market_price or c.buy_price for c in all_cards if not c.sell_price
    )

    total_value = sum(
        c.sell_price if c.sell_price else (c.market_price or c.buy_price)
        for c in all_cards
    )
    roi = ((total_value - total_invested) / total_invested * 100) if total_invested > 0 else 0

    cards_with_pl = [c for c in all_cards if c.profit_loss is not None]
    sorted_by_pl = sorted(cards_with_pl, key=lambda c: c.profit_loss, reverse=True)
    top_gainers = [c for c in sorted_by_pl if c.profit_loss > 0][:5]
    top_losers = sorted([c for c in sorted_by_pl if c.profit_loss < 0], key=lambda c: c.profit_loss)[:5]

    sport_data = defaultdict(float)
    sport_counts = defaultdict(int)
    for c in all_cards:
        sport = c.sport or 'Unknown'
        sport_counts[sport] += 1
        if c.profit_loss is not None:
            sport_data[sport] += c.profit_loss

    sport_labels = list(sport_data.keys())
    sport_profits = [round(sport_data[s], 2) for s in sport_labels]

    breakdown_labels = list(sport_counts.keys())
    breakdown_counts = [sport_counts[s] for s in breakdown_labels]

    sport_invested = defaultdict(float)
    sport_value = defaultdict(float)
    for c in all_cards:
        sport = c.sport or 'Unknown'
        sport_invested[sport] += c.buy_price
        sport_value[sport] += c.sell_price if c.sell_price else (c.market_price or c.buy_price)

    roi_labels = list(sport_invested.keys())
    roi_values = [
        round((sport_value[s] - sport_invested[s]) / sport_invested[s] * 100, 1)
        if sport_invested[s] > 0 else 0
        for s in roi_labels
    ]

    return render_template('analytics.html',
        total_cards=total_cards,
        total_invested=total_invested,
        est_market_value=est_market_value,
        realized_gains=realized_gains,
        unrealized=unrealized,
        cards_sold=cards_sold,
        roi=roi,
        top_gainers=top_gainers,
        top_losers=top_losers,
        sport_labels=sport_labels,
        sport_profits=sport_profits,
        breakdown_labels=breakdown_labels,
        breakdown_counts=breakdown_counts,
        roi_labels=roi_labels,
        roi_values=roi_values
    )

@main.route('/watchlist')
@login_required
def watchlist():
    from models import Watchlist
    items = Watchlist.query.filter_by(user_id=current_user.id).order_by(Watchlist.date_added.desc()).all()
    return render_template('watchlist.html', items=items)

@main.route('/watchlist/add', methods=['POST'])
@login_required
def add_watchlist():
    if current_user.is_demo:
        flash('Demo account cannot make changes.')
        return redirect(url_for('main.dashboard'))
    from models import Watchlist
    player_name = request.form.get('player_name', '').strip()
    if not player_name:
        flash('Player name is required.')
        return redirect(url_for('main.watchlist'))

    item = Watchlist(
        player_name=player_name,
        sport=request.form.get('sport') or None,
        target_price=float(request.form.get('target_price')) if request.form.get('target_price') else None,
        priority=request.form.get('priority', 'medium'),
        notes=request.form.get('notes', '').strip() or None,
        user_id=current_user.id
    )
    db.session.add(item)
    db.session.commit()
    flash(f'{player_name} added to watchlist.')
    return redirect(url_for('main.watchlist'))

@main.route('/watchlist/delete/<int:item_id>')
@login_required
def delete_watchlist(item_id):
    from models import Watchlist
    item = Watchlist.query.get_or_404(item_id)
    if item.user_id != current_user.id:
        flash('Permission denied.')
        return redirect(url_for('main.watchlist'))
    db.session.delete(item)
    db.session.commit()
    flash('Removed from watchlist.')
    return redirect(url_for('main.watchlist'))

@main.route('/research', methods=['GET', 'POST'])
@login_required
def research():
    player_name = None
    sport = None
    ebay_data = None
    error = None
    ai_analysis = None
    recommendation = None

    if request.method == 'POST':
        player_name = request.form.get('player_name', '').strip()
        sport = request.form.get('sport', '').strip() or None

        ebay_data, error = search_ebay_sold(player_name, sport=sport)

        if ebay_data:
            listings_text = "\n".join([
                f"- {l['title']}: ${l['price']:.2f}"
                for l in ebay_data['listings']
            ])

            prompt = f"""You are a sports card market analyst. Based on the following recent eBay listing data for {player_name} cards, provide a brief investment analysis.

eBay Data:
- Average Price: ${ebay_data['average']:.2f}
- Low: ${ebay_data['low']:.2f}
- High: ${ebay_data['high']:.2f}
- Number of Listings: {ebay_data['count']}

Recent Listings:
{listings_text}

Provide:
1. A one-word recommendation: BUY, SELL, HOLD, or WATCH
2. A 3-4 sentence analysis covering price range, market activity, and whether this player's cards represent good value right now.

Format your response as JSON like this:
{{"recommendation": "BUY", "analysis": "Your analysis here."}}"""

            try:
                client = anthropic.Anthropic(api_key=os.getenv('ANTHROPIC_API_KEY'))
                message = client.messages.create(
                    model='claude-opus-4-5',
                    max_tokens=300,
                    messages=[{'role': 'user', 'content': prompt}]
                )
                result = json.loads(message.content[0].text.strip())
                recommendation = result.get('recommendation', '')
                ai_analysis = result.get('analysis', '')
            except Exception as e:
                ai_analysis = f"AI analysis unavailable: {str(e)}"

    return render_template('research.html',
        player_name=player_name,
        sport=sport,
        ebay_data=ebay_data,
        error=error,
        ai_analysis=ai_analysis,
        recommendation=recommendation
    )

@main.route('/card_advice/<int:card_id>', methods=['GET'])
@login_required
def card_advice(card_id):
    """Get AI advice on whether to KEEP, SELL, or HOLD a specific card."""
    card = Card.query.get_or_404(card_id)

    if card.user_id != current_user.id:
        return jsonify({'error': 'Unauthorized'}), 403

    try:
        ebay_data, error = search_ebay_sold(card.player_name, sport=card.sport)

        if not ebay_data:
            return jsonify({
                'error': f'Could not fetch market data: {error or "Unknown error"}'
            }), 400

        prompt = build_ai_prompt(card, ebay_data)

        client = anthropic.Anthropic(api_key=os.getenv('ANTHROPIC_API_KEY'))
        message = client.messages.create(
            model='claude-opus-4-5',
            max_tokens=400,
            messages=[{'role': 'user', 'content': prompt}]
        )

        result = json.loads(message.content[0].text.strip())
        recommendation = result.get('recommendation', 'HOLD')
        analysis = result.get('analysis', '')

        return jsonify({
            'recommendation': recommendation,
            'analysis': analysis,
            'ebay_average': ebay_data['average'],
            'ebay_count': ebay_data['count']
        })

    except json.JSONDecodeError as e:
        return jsonify({'error': f'Failed to parse AI response: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'error': f'Error getting advice: {str(e)}'}), 500