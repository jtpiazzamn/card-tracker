from flask import Blueprint, render_template, redirect, url_for, flash, request
from flask_login import login_required, current_user
from werkzeug.security import generate_password_hash
from models import db, User, Card
from functools import wraps
import csv
import io
import openpyxl

admin = Blueprint('admin', __name__)

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash('You do not have permission to access this page.')
            return redirect(url_for('main.dashboard'))
        return f(*args, **kwargs)
    return decorated_function

@admin.route('/admin')
@login_required
@admin_required
def admin_dashboard():
    users = User.query.all()
    
    user_data = []
    for user in users:
        cards = Card.query.filter_by(user_id=user.id).all()
        total_invested = sum(card.buy_price for card in cards)
        total_value = sum(
            card.sell_price if card.sell_price
            else (card.market_price or card.buy_price)
            for card in cards
        )
        profit_loss = total_value - total_invested
        
        user_data.append({
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'is_admin': user.is_admin,
            'date_joined': user.date_joined,
            'card_count': len(cards),
            'total_invested': total_invested,
            'total_value': total_value,
            'profit_loss': profit_loss
        })
    
    total_users = len(users)
    total_cards = Card.query.count()
    
    return render_template('admin.html',
        user_data=user_data,
        total_users=total_users,
        total_cards=total_cards
    )

@admin.route('/admin/create_user', methods=['GET', 'POST'])
@login_required
@admin_required
def create_user():
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        is_admin = request.form.get('is_admin') == 'on'

        if User.query.filter_by(email=email).first():
            flash('Email already exists.')
            return redirect(url_for('admin.create_user'))

        if User.query.filter_by(username=username).first():
            flash('Username already exists.')
            return redirect(url_for('admin.create_user'))

        new_user = User(
            username=username,
            email=email,
            password=generate_password_hash(password),
            is_admin=is_admin
        )
        db.session.add(new_user)
        db.session.commit()
        flash(f'User {username} created successfully.')
        return redirect(url_for('admin.admin_dashboard'))

    return render_template('admin_create_user.html')

@admin.route('/admin/import_for_user', methods=['GET', 'POST'])
@login_required
@admin_required
def import_for_user():
    users = User.query.order_by(User.username).all()

    if request.method == 'POST':
        user_id = request.form.get('user_id')
        file = request.files.get('file')

        if not user_id or not file:
            flash('Please select a user and a file.')
            return redirect(url_for('admin.import_for_user'))

        user = User.query.get_or_404(int(user_id))
        filename = file.filename.lower()
        rows = []
        errors = []
        imported = 0

        try:
            if filename.endswith('.csv'):
                stream = io.StringIO(file.stream.read().decode('utf-8'))
                reader = csv.DictReader(stream)
                rows = list(reader)
            elif filename.endswith('.xlsx'):
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            else:
                flash('Please upload a CSV or Excel (.xlsx) file.')
                return redirect(url_for('admin.import_for_user'))

        except Exception as e:
            flash(f'Error reading file: {str(e)}')
            return redirect(url_for('admin.import_for_user'))

        for i, row in enumerate(rows, start=2):
            try:
                player_name = row.get('Player Name') or row.get('player_name') or row.get('Name')
                if not player_name:
                    errors.append(f'Row {i}: Missing player name')
                    continue

                buy_price_raw = row.get('Buy Price') or row.get('buy_price') or '0'
                buy_price_str = str(buy_price_raw).replace('$', '').replace(',', '').strip()
                buy_price = float(buy_price_str) if buy_price_str else 0.0

                sell_price_raw = row.get('Sell Price') or row.get('sell_price') or ''
                sell_price_str = str(sell_price_raw).replace('$', '').replace(',', '').strip()
                sell_price = float(sell_price_str) if sell_price_str else None

                market_price_raw = row.get('Market Price') or row.get('market_price') or ''
                market_price_str = str(market_price_raw).replace('$', '').replace(',', '').strip()
                market_price = float(market_price_str) if market_price_str else None

                new_card = Card(
                    player_name=str(player_name).strip(),
                    sport=str(row.get('Sport') or row.get('sport') or '').strip() or None,
                    year=str(row.get('Year') or row.get('year') or '').strip() or None,
                    manufacturer=str(row.get('Manufacturer') or row.get('manufacturer') or '').strip() or None,
                    condition=str(row.get('Condition') or row.get('condition') or '').strip() or None,
                    buy_price=buy_price,
                    sell_price=sell_price,
                    market_price=market_price,
                    user_id=user.id
                )
                db.session.add(new_card)
                imported += 1

            except Exception as e:
                errors.append(f'Row {i}: {str(e)}')

        db.session.commit()

        if errors:
            flash(f'Imported {imported} cards for {user.username}. {len(errors)} rows had errors: {", ".join(errors[:3])}')
        else:
            flash(f'Successfully imported {imported} cards for {user.username}.')

        return redirect(url_for('admin.admin_dashboard'))

    return render_template('admin_import_user.html', users=users)

@admin.route('/admin/make_admin/<int:user_id>')
@login_required
@admin_required
def make_admin(user_id):
    user = User.query.get_or_404(user_id)
    user.is_admin = True
    db.session.commit()
    flash(f'{user.username} is now an admin.')
    return redirect(url_for('admin.admin_dashboard'))

@admin.route('/admin/remove_admin/<int:user_id>')
@login_required
@admin_required
def remove_admin(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash('You cannot remove your own admin status.')
        return redirect(url_for('admin.admin_dashboard'))
    user.is_admin = False
    db.session.commit()
    flash(f'{user.username} admin status removed.')
    return redirect(url_for('admin.admin_dashboard'))
@admin.route('/admin/reset_password/<int:user_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def reset_password(user_id):
    user = User.query.get_or_404(user_id)
    if request.method == 'POST':
        new_password = request.form.get('new_password')
        user.password = generate_password_hash(new_password)
        db.session.commit()
        flash(f'Password reset for {user.username}.')
        return redirect(url_for('admin.admin_dashboard'))
    return render_template('admin_reset_password.html', user=user)
@admin.route('/admin/delete_user/<int:user_id>')
@login_required
@admin_required
def delete_user(user_id):
    if user_id == current_user.id:
        flash('You cannot delete your own account.')
        return redirect(url_for('admin.admin_dashboard'))
    user = User.query.get_or_404(user_id)
    Card.query.filter_by(user_id=user_id).delete()
    db.session.delete(user)
    db.session.commit()
    flash(f'User {user.username} has been deleted.')
    return redirect(url_for('admin.admin_dashboard'))
